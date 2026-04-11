import os
import csv
import io
from datetime import datetime, timezone
from flask import current_app

from ..extensions import db
from ..models.report import Report, ReportFormat
from ..models.audit_log import AuditLog
from ..models.user import User


def _reports_dir() -> str:
    d = current_app.config.get("REPORTS_DIR", "reports")
    os.makedirs(d, exist_ok=True)
    return d


def _timestamp_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


# ── PDF ────────────────────────────────────────────────────────────────────────

def _generate_pdf(title: str, headers: list, rows: list, file_path: str) -> None:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 12))

    table_data = [headers] + [[str(cell) if cell is not None else "" for cell in row] for row in rows]
    col_count = len(headers)
    col_width = (landscape(A4)[0] - 72) / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elements.append(table)
    doc.build(elements)


# ── Excel ──────────────────────────────────────────────────────────────────────

def _generate_excel(title: str, headers: list, rows: list, file_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(file_path)


# ── CSV ────────────────────────────────────────────────────────────────────────

def _generate_csv(headers: list, rows: list, file_path: str) -> None:
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


# ── Data builders ──────────────────────────────────────────────────────────────

def _build_audit_log_data(params: dict) -> tuple[list, list]:
    query = AuditLog.query.order_by(AuditLog.timestamp.desc())

    if params.get("start_date"):
        query = query.filter(AuditLog.timestamp >= params["start_date"])
    if params.get("end_date"):
        query = query.filter(AuditLog.timestamp <= params["end_date"])
    if params.get("user_id"):
        query = query.filter(AuditLog.user_id == params["user_id"])
    if params.get("action"):
        query = query.filter(AuditLog.action.ilike(f"%{params['action']}%"))

    limit = int(params.get("limit", 1000))
    logs = query.limit(limit).all()

    headers = ["ID", "Timestamp", "User", "Action", "Resource", "Status", "IP Address"]
    rows = [
        [
            log.id,
            log.timestamp.isoformat() if log.timestamp else "",
            log.user.username if log.user else f"user_id={log.user_id}",
            log.action,
            log.resource or "",
            log.status.value,
            log.ip_address or "",
        ]
        for log in logs
    ]
    return headers, rows


def _build_users_data(params: dict) -> tuple[list, list]:
    users = User.query.order_by(User.created_at.desc()).all()
    headers = ["ID", "Username", "Email", "Role", "Active", "Approved", "Created At", "Last Login"]
    rows = [
        [
            u.id,
            u.username,
            u.email,
            u.role.value,
            u.is_active,
            u.is_approved,
            u.created_at.isoformat() if u.created_at else "",
            u.last_login.isoformat() if u.last_login else "",
        ]
        for u in users
    ]
    return headers, rows


# ── Public API ─────────────────────────────────────────────────────────────────

REPORT_BUILDERS = {
    "audit_logs": _build_audit_log_data,
    "users": _build_users_data,
}


def generate_report(
    title: str,
    report_type: str,
    fmt: str,
    generated_by_id: int,
    parameters: dict = None,
) -> Report:
    parameters = parameters or {}

    if report_type not in REPORT_BUILDERS:
        raise ValueError(f"Unknown report_type '{report_type}'. Valid: {list(REPORT_BUILDERS)}")

    if fmt not in (f.value for f in ReportFormat):
        raise ValueError(f"Unknown format '{fmt}'.")

    headers, rows = REPORT_BUILDERS[report_type](parameters)

    ts = _timestamp_str()
    extension = {"pdf": "pdf", "excel": "xlsx", "csv": "csv"}[fmt]
    # Filename is built entirely from whitelisted values + timestamp — safe
    filename = f"{report_type}_{ts}.{extension}"
    reports_dir = _reports_dir()
    file_path = os.path.join(os.path.realpath(reports_dir), filename)

    if fmt == "pdf":
        _generate_pdf(title, headers, rows, file_path)
    elif fmt == "excel":
        _generate_excel(title, headers, rows, file_path)
    else:
        _generate_csv(headers, rows, file_path)

    report = Report(
        title=title,
        report_type=report_type,
        generated_by=generated_by_id,
        format=ReportFormat(fmt),
        file_path=file_path,
        parameters=parameters,
    )
    db.session.add(report)
    db.session.commit()
    return report
